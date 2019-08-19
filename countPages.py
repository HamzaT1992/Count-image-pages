import os
import openpyxl as xl
from PIL import Image
from PyPDF2 import PdfFileReader

#wb=xl.Workbook()
wb = xl.load_workbook('tbvue.xlsx')
ws = wb.active
for i in range(2,3860):
    filename, ext = os.path.splitext(ws['Q'+str(i)].value)
    nbValue = ws['W'+str(i)].value
    if ext == '.tif' and nbValue == 'NULL':
        img = Image.open(ws['Q'+str(i)].value)
        ws['W'+str(i)] = img.n_frames
    elif ext == '.pdf' and nbValue == 'NULL':
        pdf = PdfFileReader(open(ws['Q'+str(i)].value,'rb'))
        ws['W'+str(i)] = pdf.getNumPages()
wb.save('tbvue2.xlsx')